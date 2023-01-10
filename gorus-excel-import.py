import os
from urllib import request

import openpyxl
from django.core.asgi import get_asgi_application

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
application = get_asgi_application()

from apps.training.models import Community, Interview


def column_index_value(cname, max_col, sheet_obj):
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row=1, column=i)
        if cell_obj.value == cname:
            return i
    return -1


def excel_read(request):
    wb_obj = openpyxl.load_workbook('gorusmeler.xlsx')
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column

    kod_index = column_index_value('kod', max_col, sheet_obj)
    durum_index = column_index_value('durum', max_col, sheet_obj)
    not_index = column_index_value('not', max_col, sheet_obj)

    for r in range(2, sheet_obj.max_row + 1):
        try:
            kod_value = str(sheet_obj.cell(row=r, column=kod_index).value)
            durum_value = int(sheet_obj.cell(row=r, column=durum_index).value)
            not_value = str(sheet_obj.cell(row=r, column=not_index).value)

            community = Community.objects.filter(period_code=kod_value).first()

            if durum_value:
                control = Interview.objects.create(
                    community=community,
                    user_id=4,
                    note=not_value,
                    interview_status_id=durum_value,
                )
            else:
                control = Interview.objects.create(
                    community=community,
                    user_id=4,
                    note=not_value,
                )

            if control:
                print('Yükleme Tamamlandı')
            else:
                print('Yükleme Hatası id:' + str(kod_value))

        except Exception as ex:
            print(ex)


excel_read(request)
