# encoding:utf-8

import datetime

import requests
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.utils.translation import gettext_lazy as _

from apps.mainpage.views import wrong404
from apps.profile.forms import ProfileForm, UserForm, ContactForm
from apps.profile.models import Profile
from apps.training.forms import InterviewForm
from apps.training.models import *
import openpyxl


def capitalize(incoming_text):
    metin = incoming_text.replace("i", "İ")
    metin = metin.upper()
    return metin


@login_required
def community(request):
    page_control = request.GET.get('page')
    if not page_control:
        return HttpResponseRedirect('/training/community/' + '?page=1')
    community_get_id = request.GET.get('community_get_id')
    communities = Community.objects.filter().order_by('full_name')
    interviews_statu_list = InterviewStatus.objects.filter()
    paginator = Paginator(communities, 25)
    page = request.GET.get('page', 1)
    try:
        communities_list = paginator.page(page)
    except PageNotAnInteger:
        communities_list = paginator.page(1)
    except EmptyPage:
        communities_list = paginator.page(paginator.num_pages)
    context = {
        'communities': communities_list,
        'interviews_statu_list': interviews_statu_list,
        'community_get_id': community_get_id,
        'page': page,
    }
    return render(request, "apps/training/communities.html", context)


@login_required
def community_update(request, id):
    if request.POST:
        block1 = request.POST.get('add_block')
        entrance1 = request.POST.get('add_entrance')
        apartment1 = request.POST.get('add_apartment')
        period1 = request.POST.get('add_period')
        period_code1 = request.POST.get('add_period_code')
        m21 = request.POST.get('add_m2')
        full_name1 = request.POST.get('add_full_name')
        full_name1 = capitalize(full_name1)
        phone_one1 = request.POST.get('add_phone_one')
        phone_two1 = request.POST.get('add_phone_two')
        phone_three1 = request.POST.get('add_phone_three')
        tc1 = request.POST.get('add_tc')
        circuit_start1 = request.POST.get('add_circuit_start')
        circuit_finish1 = request.POST.get('add_circuit_finish')
        note1 = request.POST.get('add_note')
        interview_status1 = request.POST.get('add_interview_status_test')
        control = Community.objects.filter(id=id).update(
            block=block1,
            entrance=entrance1,
            apartment=apartment1,
            period=period1,
            period_code=period_code1,
            m2=m21,
            full_name=full_name1,
            phone_one=phone_one1,
            phone_two=phone_two1,
            phone_three=phone_three1,
            tc=tc1,
            circuit_start=circuit_start1,
            circuit_finish=circuit_finish1,
        )

        control_two = Interview.objects.create(
            community_id=id,
            user=request.user,
            note=note1,
            interview_status_id=interview_status1,
        )

        messages.success(request, _('Güncelleme İşlemi Başarılı'))
    else:
        messages.warning(request, _('İzinsiz Giriş'))

    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))


@login_required
def updateInterview(request, id):
    if request.POST:
        note = request.POST.get('add_note')
        interview_status_id = request.POST.get('add_interview_status_test')

        price_sales = request.POST.get('add_price_sales')
        price_sales = price_sales.split(",")
        price_sales = price_sales[0]

        price_rent = request.POST.get('add_price_rent')
        price_rent = price_rent.split(",")
        price_rent = price_rent[0]

        control_two = Interview.objects.filter(id=id).update(
            user=request.user,
            note=note,
            interview_status_id=interview_status_id,
            price_sales=price_sales,
            price_rent=price_rent,
        )

        messages.success(request, _('Güncelleme İşlemi Başarılı'))
    else:
        messages.warning(request, _('İzinsiz Giriş'))

    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))


@login_required
def community_add(request):
    if request.POST:
        block = request.POST.get('add_block')
        entrance = request.POST.get('add_entrance')
        apartment = request.POST.get('add_apartment')
        period = request.POST.get('add_period')
        period_code = request.POST.get('add_period_code')
        m2 = request.POST.get('add_m2')
        full_name = request.POST.get('add_full_name')
        full_name = capitalize(full_name)
        phone_one = request.POST.get('add_phone_one')
        phone_two = request.POST.get('add_phone_two')
        phone_three = request.POST.get('add_phone_three')
        tc = request.POST.get('add_tc')
        circuit_start = request.POST.get('add_circuit_start')
        circuit_finish = request.POST.get('add_circuit_finish')

        note = request.POST.get('add_note')
        interview_status = request.POST.get('add_interview_status')
        price_sales = request.POST.get('add_price_sales')
        price_rent = request.POST.get('add_price_rent')

        control = Community.objects.create(
            block=block,
            entrance=entrance,
            apartment=apartment,
            period=period,
            period_code=period_code,
            m2=m2,
            full_name=full_name,
            phone_one=phone_one,
            phone_two=phone_two,
            phone_three=phone_three,
            tc=tc,
            circuit_start=circuit_start,
            circuit_finish=circuit_finish,
        )

        control_two = Interview.objects.create(
            community=control,
            user=request.user,
            note=note,
            interview_status_id=interview_status,
            price_sales=price_sales,
            price_rent=price_rent,
        )

        messages.success(request, _('Kayıt İşlemi Başarılı'))
    else:
        messages.warning(request, _('İzinsiz Giriş'))
    return redirect('training:community')


@login_required
def deleteCommunity(request, id):
    community = Community.objects.get(id=id)
    community.delete()
    messages.success(request, _('Silme İşlemi Başarılı'))
    return redirect('training:community')


@login_required
def interviews(request):
    interviews_statu_list = InterviewStatus.objects.filter()
    interviews = Interview.objects.filter()
    paginator = Paginator(interviews, 25)
    page = request.GET.get('page', 1)
    try:
        interviews_list = paginator.page(page)
    except PageNotAnInteger:
        interviews_list = paginator.page(1)
    except EmptyPage:
        interviews_list = paginator.page(paginator.num_pages)
    context = {
        'interviews_list': interviews_list,
        'interviews_statu_list': interviews_statu_list
    }
    return render(request, "apps/training/interviews.html", context)


@login_required
def detailInterview(request, id):
    community_detail = Community.objects.get(id=id)
    interviews_statu_list = InterviewStatus.objects.filter()
    context = {
        'community': community_detail,
        'interviews_statu_list': interviews_statu_list,
    }
    return render(request, "apps/training/interview_detail.html", context)


@login_required
def addInterview(request, id):
    if request.method == "POST":
        interview_status = request.POST.get('interview_status')
        price_sales = request.POST.get('price_sales')
        price_rent = request.POST.get('price_rent')
        note = request.POST.get('note')
        if interview_status and price_sales and price_rent and note:
            control = Interview.objects.create(
                community_id=id,
                user=request.user,
                note=note,
                interview_status_id=interview_status,
                price_sales=price_sales,
                price_rent=price_rent,
            )
            if control:
                messages.success(request, "Görüşme Başarıyla Kaydedildi")
            else:
                messages.warning(request, "Beklenmedik Bir Hata Oluştu")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
        else:
            messages.warning(request, "Formu eksiksiz doldurduğunuza emin olun")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))


@login_required
def interviewDelete(request, id):
    interview = Interview.objects.get(id=id)
    interview.delete()
    messages.success(request, _('Silme İşlemi Başarılı'))
    return redirect('training:interviews')


@login_required
def interviews_statu(request):
    interviews_statu_list = InterviewStatus.objects.all()
    context = {'interviews_statu_list': interviews_statu_list}
    return render(request, "apps/training/interviews_statu.html", context)


@login_required
def interviews_statu_add(request):
    if request.method == "POST":
        text = request.POST.get('add_text')
        if text:
            control = InterviewStatus.objects.create(
                text=text,
            )
            if control:
                messages.success(request, "Durum Başarıyla Kaydedildi")
            else:
                messages.warning(request, "Beklenmedik Bir Hata Oluştu")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
        else:
            messages.warning(request, "Formu eksiksiz doldurduğunuza emin olun")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))


@login_required
def interviews_statu_delete(request, id):
    interviews_statu_list = InterviewStatus.objects.all()
    context = {'interviews_statu_list': interviews_statu_list}
    return render(request, "apps/training/interviews_statu.html", context)


@login_required
def interviews_statu_update(request, id):
    interviews_statu_list = InterviewStatus.objects.all()
    context = {'interviews_statu_list': interviews_statu_list}
    return render(request, "apps/training/interviews_statu.html", context)


@login_required
def search(request):
    interviews_statu_list = InterviewStatus.objects.filter()
    community_block_list = Community.objects.order_by('block').values_list('block', flat=True).distinct()
    community_entrance_list = Community.objects.order_by('entrance').values_list('entrance', flat=True).distinct()
    community_apartment_list = Community.objects.order_by('apartment').values_list('apartment', flat=True).distinct()
    community_period_list = Community.objects.order_by('period').values_list('period', flat=True).distinct()
    community_m2_list = Community.objects.order_by('m2').values_list('m2', flat=True).distinct()
    last_list = []
    id_list = []
    if request.POST:
        full_name = request.POST.get('full_name')
        if full_name:
            full_name = capitalize(full_name)
        tc = request.POST.get('tc')
        phone = request.POST.get('phone')
        period_code = request.POST.get('period_code')

        block = request.POST.getlist('block')
        entrance = request.POST.getlist('entrance')
        apartment = request.POST.getlist('apartment')
        period = request.POST.getlist('period')
        m2 = request.POST.getlist('m2')
        interview_status_id = request.POST.getlist('interview_status_id', 0)

        note_2 = request.POST.get('note_2')

        filters = {
            'period_code__isnull': False
        }
        if full_name:
            filters['full_name__contains'] = full_name
        if tc:
            filters['tc'] = tc
        if phone:
            filters['phone_one'] = phone
        if period_code:
            filters['period_code'] = period_code
        if block:
            filters['block__in'] = block
        if entrance:
            filters['entrance__in'] = entrance
        if apartment:
            filters['apartment__in'] = apartment
        if period:
            filters['period__in'] = period
        if m2:
            filters['m2__in'] = m2
        if interview_status_id:
            for interview_id in interview_status_id:
                interview_status = InterviewStatus.objects.filter(id=interview_id).last()
                id_list.append(interview_status)
            filters['interview_community__interview_status__in'] = id_list
        if note_2:
            filters['interview_community__note__contains'] = note_2
        communities_all_list = Community.objects.filter(**filters)
        if interview_status_id:
            for communities_item in communities_all_list:
                if communities_item.interview_last().interview_status in id_list:
                    last_list.append(communities_item)
        else:
            last_list = communities_all_list
        context = {
            'communities': last_list,
            'interviews_statu_list': interviews_statu_list,
            'community_block_list': community_block_list,
            'community_entrance_list': community_entrance_list,
            'community_apartment_list': community_apartment_list,
            'community_period_list': community_period_list,
            'community_m2_list': community_m2_list,
        }
    else:
        context = {
            'interviews_statu_list': interviews_statu_list,
            'community_block_list': community_block_list,
            'community_entrance_list': community_entrance_list,
            'community_apartment_list': community_apartment_list,
            'community_period_list': community_period_list,
            'community_m2_list': community_m2_list,
        }
    return render(request, "apps/training/search.html", context)


def column_index_value(cname, max_col, sheet_obj):
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row=1, column=i)
        if cell_obj.value == cname:
            return i
    return -1


@login_required
def data_transfer(request):
    if request.method == 'POST':
        community_list = []
        file = request.FILES['file-upload']

        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column

        id_index = column_index_value('id', max_col, sheet_obj)
        blok_index = column_index_value('blok', max_col, sheet_obj)
        giris_index = column_index_value('giris', max_col, sheet_obj)
        daire_index = column_index_value('daire', max_col, sheet_obj)
        donem_index = column_index_value('donem', max_col, sheet_obj)
        donem_kodu_index = column_index_value('donem_kodu', max_col, sheet_obj)
        m2_index = column_index_value('m2', max_col, sheet_obj)
        name_index = column_index_value('name', max_col, sheet_obj)
        telefon_1_index = column_index_value('telefon_1', max_col, sheet_obj)
        telefon_2_index = column_index_value('telefon_2', max_col, sheet_obj)
        telefon_3_index = column_index_value('telefon_3', max_col, sheet_obj)
        tc_index = column_index_value('tc', max_col, sheet_obj)
        tarih_1_index = column_index_value('tarih_1', max_col, sheet_obj)
        tarih_2_index = column_index_value('tarih_2', max_col, sheet_obj)
        for r in range(2, sheet_obj.max_row + 1):
            id_value = int(sheet_obj.cell(row=r, column=id_index).value)
            blok_value = str(sheet_obj.cell(row=r, column=blok_index).value)
            giris_value = str(sheet_obj.cell(row=r, column=giris_index).value)
            daire_value = str(sheet_obj.cell(row=r, column=daire_index).value)
            donem_value = str(sheet_obj.cell(row=r, column=donem_index).value)
            donem_kodu_value = str(sheet_obj.cell(row=r, column=donem_kodu_index).value)
            m2_value = str(sheet_obj.cell(row=r, column=m2_index).value)
            name_value = str(sheet_obj.cell(row=r, column=name_index).value)
            name_value = capitalize(name_value)
            telefon_1_value = str(sheet_obj.cell(row=r, column=telefon_1_index).value)
            telefon_2_value = str(sheet_obj.cell(row=r, column=telefon_2_index).value)
            telefon_3_value = str(sheet_obj.cell(row=r, column=telefon_3_index).value)
            tc_value = str(sheet_obj.cell(row=r, column=tc_index).value)
            tarih_1_value = str(sheet_obj.cell(row=r, column=tarih_1_index).value)
            tarih_2_value = str(sheet_obj.cell(row=r, column=tarih_2_index).value)
            main_control = Community.objects.filter(period_code=donem_kodu_value).last()
            if main_control:
                control = Community.objects.filter(period_code=donem_kodu_value).update(
                    block=blok_value,
                    entrance=giris_value,
                    apartment=daire_value,
                    period=donem_value,
                    period_code=donem_kodu_value,
                    m2=m2_value,
                    full_name=name_value,
                    phone_one=telefon_1_value,
                    phone_two=telefon_2_value,
                    phone_three=telefon_3_value,
                    tc=tc_value,
                    circuit_start=tarih_1_value,
                    circuit_finish=tarih_2_value,
                )
            else:
                control = Community.objects.create(
                    block=blok_value,
                    entrance=giris_value,
                    apartment=daire_value,
                    period=donem_value,
                    period_code=donem_kodu_value,
                    m2=m2_value,
                    full_name=name_value,
                    phone_one=telefon_1_value,
                    phone_two=telefon_2_value,
                    phone_three=telefon_3_value,
                    tc=tc_value,
                    circuit_start=tarih_1_value,
                    circuit_finish=tarih_2_value,
                )
            if control:
                print('Yükleme Tamamlandı')
            else:
                print('Yükleme Hatası id:' + str(id_value))

            messages.success(request, "Veriler Başarıyla Aktarıldı")
            community_list.append(control)
        context = {'community_list': community_list}
    else:
        context = {}
    return render(request, "apps/training/data_transfer.html", context)
